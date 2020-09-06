import openpyxl
import csv
import pandas as pd
import Levenshtein
import copy

# open the workbook
wb = openpyxl.load_workbook('NARA NY CECF DB check list 2020 Summer.xlsx')
# open the sheet
sheet = wb['CHINESE']

wb_street = openpyxl.load_workbook('street.xlsx')
sheet_street = wb_street['STREET']


def get_col_number(sheet, name):
    for col in range(1, sheet.max_column + 1):
        if sheet.cell(1, col).value == name:
            return col


def get_data_list(sheet, col):
    lst = []
    for row in range(2, sheet.max_row + 1):
        lst.append(sheet.cell(row, col).value)
    return lst


# get all entry date infor into a list
def clean_entry_date(sheet):
    col = get_col_number(sheet, 'ENTRYDATE')
    lst = get_data_list(sheet, col)

    years = []
    months = []
    days = []
    for i in range(len(lst)):
        value = lst[i]
        if type(value) == type('1'):
            if len(value) == 8:
                year = value[-4:]
                month = value[0:2]
                if value[2:4] == '  ':
                    day = None
                else:
                    day = value[2:4]
            elif len(value) == 4:
                year = value
                month = None
                day = None
        else:
            year = None
            month = None
            day = None
        years.append(year)
        months.append(month)
        days.append(day)

    df = pd.read_csv("you_xu.csv")
    df["ENTRYDATE_YEAR"] = years
    df['ENTRYDATE_MONTH'] = months
    df['ENTRYDATE_DAY'] = days
    df.to_csv("you_xu.csv", index=False)


def clean_entry_date2(sheet):
    col = get_col_number(sheet, 'ENTRYDATE2')
    lst = get_data_list(sheet, col)

    years = []
    months = []
    days = []
    for i in range(len(lst)):
        value = lst[i]
        if type(value) == type('1'):
            if len(value) == 8:
                year = value[-4:]
                month = value[0:2]
                if value[2:4] == '  ':
                    day = None
                else:
                    day = value[2:4]
            elif len(value) == 4:
                year = value
                month = None
                day = None
        else:
            year = None
            month = None
            day = None
        years.append(year)
        months.append(month)
        days.append(day)

    df = pd.read_csv("you_xu.csv")
    df["ENTRYDATE2_YEAR"] = years
    df['ENTRYDATE2_MONTH'] = months
    df['ENTRYDATE2_DAY'] = days
    df.to_csv("you_xu.csv", index=False)


def clean_docdate(sheet):
    col = get_col_number(sheet, 'DOCDATE')
    lst = get_data_list(sheet, col)

    first = []
    last = []
    for date in lst:
        if type(date) == type('1') and len(date) == 8:
            first.append(date[:4])
            last.append(date[-4:])
        else:
            first.append(None)
            last.append(None)

    df = pd.read_csv("you_xu.csv")
    df["DOCDATE1"] = first
    df['DOCDATE2'] = last
    df.to_csv("you_xu.csv", index=False)


def clean_birthplace(sheet):
    col = get_col_number(sheet, 'BIRTHPLACE')
    lst = get_data_list(sheet, col)
    #print(len(lst))

    country = []
    state = []
    city = []
    for data in lst:
        if data == None:
            country.append('')
            state.append('')
            city.append('')
        else:
            placelist = [i.strip() for i in data.split(',')]
            if len(placelist) == 1:
                country.append(placelist[0])
                state.append('')
                city.append('')
            elif len(placelist) == 2:
                if len(placelist[1]) == 2:
                    placelist[0] = placelist[0].replace('?', '')
                    placelist[1] = placelist[1].replace('?', '')
                    country.append('U.S.')
                    state.append(placelist[1])
                    city.append(placelist[0])
                else:
                    placelist[0] = placelist[0].replace('?', '')
                    placelist[1] = placelist[1].replace('?', '')
                    country.append(placelist[1])
                    state.append(placelist[0])
                    city.append('')
            elif len(placelist) == 3:
                placelist[0] = placelist[0].replace('?', '')
                placelist[1] = placelist[1].replace('?', '')
                placelist[2] = placelist[2].replace('?', '')
                country.append(placelist[2])
                state.append(placelist[1])
                city.append(placelist[0])
    #print(len(country))
    #print(len(state))
    #print(len(city))

    df = pd.read_csv("you_xu.csv")
    df["BIRTHPLACE_CITY/COUNTY"] = city
    df['BIRTHPLACE_STATE/CITY'] = state
    df['BIRTHPLACE_COUNTRY/REGION'] = country
    df.to_csv("you_xu.csv", index=False)


def clean_street2(sheet):
    col = get_col_number(sheet, 'STREET2')
    lst = get_data_list(sheet, col)

    number = []
    strt = []
    tp = []
    rm = []

    for data in lst:
        if data == None:
            number.append(None)
            strt.append(None)
            tp.append(None)
            rm.append(None)
        else:
            data = data.replace('.', '')
            data1 = [i.strip() for i in data.split(',')]
            if len(data1) == 2:
                rm.append(data1[1])
            else:
                rm.append(None)
            for i in range(len(data1[0])):
                if data1[0][i] == " ":
                    index1 = i
                    break
            for j in range(len(data1[0])-1, -1, -1):
                if data1[0][j] == ' ':
                    index2 = j
                    break
            number.append(data1[0][0:index1].replace('?', ""))
            strt.append(data1[0][index1 + 1: index2].replace('?', ''))
            tp.append(data1[0][index2 + 1:].replace('?', ''))

    df = pd.read_csv("you_xu.csv")
    df["STREET2_NUMBER"] = number
    df['STREET2_STREET'] = strt
    df['STREET2_TYPE'] = tp
    df['STREET2_ROOM'] = rm
    df.to_csv("you_xu.csv", index=False)

def clean_hometown(sheet):
    col = get_col_number(sheet, 'HOMETOWN')
    lst = get_data_list(sheet, col)
    print(len(lst))
    final = [None] * len(lst)
    remain = [i for i in range(len(lst))]

    while len(remain) > 0:
        remain_2 = []
        similar_index = []
        similar = []
        current_index = 0
        while lst[remain[current_index]] is None:
            current_index += 1
        current = lst[remain[current_index]]

        similar.append(current)
        similar_index.append(remain[current_index])
        for j in range(current_index + 1, len(remain)):
            if lst[remain[j]] is not None:
                other = lst[remain[j]]
                other_index = remain[j]
                #print(current, other, Levenshtein.ratio(current, other))
                if Levenshtein.ratio(current, other) > 0.95:
                    similar.append(other)
                    similar_index.append(other_index)
                else:
                    remain_2.append(other_index)
        median = Levenshtein.median(similar)
        for index in similar_index:
            final[index] = median
        remain = copy.deepcopy(remain_2)

    print(len(final))
    df = pd.read_csv("you_xu.csv")
    df["HOMETOWN_NEW"] = final
    df.to_csv("you_xu.csv", index=False)

if __name__ == '__main__':
    # clean_entry_date(sheet)
    # clean_entry_date2(sheet)
    # clean_docdate(sheet)
    # clean_birthplace(sheet)
    # clean_street2(sheet_street)
    clean_hometown(sheet_street)
